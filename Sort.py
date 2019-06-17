import xlrd as xl
import openpyxl as ox
class ReadWrite:
    def __init__(self, loc_r, loc_w):
        self.read_location = loc_r
        self.write_location = loc_w
        self.row_index = 1
        self.sheet_name = ""
        self.num_cols = 0
        self.num_sheets = 0
        print("init")

    def read_write(self):
        # open read workbook
        wb_read = xl.open_workbook(self.read_location)
        sheet_r = wb_read.sheet_by_index(0)
        print("open")
        # initialize sheet name
        self.sheet_name = sheet_r.cell_value(1, 2)
        print(self.sheet_name)

        self.num_cols = sheet_r.ncols

        # open write workbook
        wb = ox.Workbook()
        sheet_w = wb.active
        wb.create_sheet(index=1, title=self.sheet_name)

        """

        REQUIRED FORMAT:
        Each sheet has 1 row of headers, at index 1
        Note that indexing starts at 1 not 0
        Sheets in output file will be grouped according to the 3rd row of read file

        DATATYPES:
        row is an array of strings. It stores all values in a single row


        """

        # Loop through every row, directly write cell values

        for r in range(sheet_r.nrows):
            row = []
            for c in range(self.num_cols):
                row.append(sheet_r.cell_value(r, c))

            print(self.row_index)

            if sheet_r.cell_value(r, 2) != self.sheet_name:
                # change sheet name, create new sheet, start index from 0
                self.sheet_name = sheet_r.cell_value(r, 2)
                print(self.sheet_name)
                self.num_sheets += 1
                wb.create_sheet(index=self.num_sheets, title=self.sheet_name)
                self.row_index = 0

            # if a new label is encountered, create new sheet with said label

            for c in range(self.num_cols):
                temp = sheet_w.cell(self.row_index + 1, c + 1)
                temp.value = row[c]

                # save sheet every time
                wb.save(self.write_location)

            self.row_index += 1

        # end of for r in range loop

    # END OF read_write()


def main():

    print("start")

    sort = ReadWrite("C:\\Users\hanissa.shamsuddin\Work Folders\VS2017 Repository\Comparison\Strictly Filtered - Final.xlsx",
                     "C:\\Users\hanissa.shamsuddin\Work Folders\VS2017 Repository\Comparison\Output.xlsx")

    sort.read_write()

    print("Done")

main()